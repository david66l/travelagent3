"""OutputFormatAgent — Markdown polish, PDF/Excel export, map links.

Produces multi-modal itinerary artifacts from a planner-generated itinerary.
Failures in any formatter are isolated: the agent always returns at least the
original Markdown text.
"""

from __future__ import annotations

import logging
import os
import platform
import re
import uuid
from datetime import datetime
from typing import Any

from core.settings import settings

logger = logging.getLogger(__name__)

# Output directories relative to project root.
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
PDF_DIR = os.path.join(_PROJECT_ROOT, "outputs", "pdfs")
EXCEL_DIR = os.path.join(_PROJECT_ROOT, "outputs", "excel")


def _ensure_dirs() -> None:
    os.makedirs(PDF_DIR, exist_ok=True)
    os.makedirs(EXCEL_DIR, exist_ok=True)


def _safe_filename(prefix: str, ext: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}.{ext}"


def _strip_markdown_for_excel(text: str) -> str:
    """Remove markdown syntax for plain-cell export."""
    text = re.sub(r"#+\s*", "", text)
    text = re.sub(r"\*\*|__", "", text)
    text = re.sub(r"`", "", text)
    text = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", text)
    return text.strip()


class OutputFormatAgent:
    """Format itinerary output into Markdown, PDF, Excel and map links."""

    def __init__(self, base_url: str | None = None) -> None:
        if base_url:
            self.base_url = base_url
        elif settings.public_app_url:
            self.base_url = settings.public_app_url.rstrip("/")
        else:
            # 0.0.0.0 is fine for binding but not for client-side download URLs.
            host = "localhost" if settings.app_host == "0.0.0.0" else settings.app_host
            self.base_url = f"http://{host}:{settings.app_port}"
        _ensure_dirs()

    async def format(
        self,
        proposal_text: str,
        itinerary: list[dict[str, Any]] | None = None,
        city: str | None = None,
        session_id: str | None = None,
        on_token: Any = None,
    ) -> dict[str, Any]:
        """Return formatted outputs and download URLs.

        Always returns at least `markdown`. PDF/Excel/map_url are best-effort.
        ``on_token`` (async ``(chunk) -> None``) streams the polish token-by-token
        to the caller (→ frontend) so the prose appears live instead of after a
        ~2k-token blocking call.
        """
        polished = await self.polish_markdown(proposal_text, on_token=on_token)
        file_id = session_id or uuid.uuid4().hex[:12]

        # Server-side PDF/Excel is redundant with the frontend's client-side
        # export, so it is gated off by default (saves CPU + the WeasyPrint native
        # dependency). When disabled, the URLs are simply omitted.
        pdf_path = pdf_url = excel_path = excel_url = None
        if settings.server_side_export_enabled:
            import asyncio

            pdf_task = asyncio.create_task(self.generate_pdf(polished, file_id))
            excel_task = asyncio.create_task(
                self.generate_excel(polished, itinerary or [], file_id)
            )
            (pdf_path, pdf_url), (excel_path, excel_url) = await asyncio.gather(
                pdf_task, excel_task
            )
        map_url = self.generate_map_url(itinerary, city)

        return {
            "markdown": polished,
            "output_markdown": polished,
            "output_pdf_url": pdf_url,
            "output_excel_url": excel_url,
            "output_map_url": map_url,
            "files": {
                "pdf_path": pdf_path,
                "excel_path": excel_path,
            },
        }

    @staticmethod
    def _build_skeleton(
        itinerary: list[dict[str, Any]] | None, profile: dict[str, Any] | None = None
    ) -> str:
        """Compact, factual text of the solved itinerary for the prose prompt."""
        if not itinerary:
            return ""
        profile = profile or {}
        dest = profile.get("destination") or "目的地"
        days = len(itinerary)
        lines = [f"目的地：{dest}；行程天数：{days}天"]
        for day in itinerary:
            lines.append(f"\n第{day.get('day_number', '?')}天：")
            for act in day.get("activities", []) or []:
                cost = act.get("ticket_price") or act.get("meal_cost") or 0
                time_str = (
                    f"{act.get('start_time', '')}-{act.get('end_time', '')}"
                    if act.get("start_time")
                    else ""
                )
                tags = "、".join(act.get("tags") or [])
                cost_str = f"，门票¥{cost:.0f}" if cost else ""
                lines.append(
                    f"- {time_str} {act.get('poi_name', '')}"
                    f"（{act.get('category', '')}{cost_str}）" + (f" 标签：{tags}" if tags else "")
                )
        return "\n".join(lines)

    async def stream_markdown(
        self,
        itinerary: list[dict[str, Any]] | None,
        profile: dict[str, Any] | None = None,
        on_token: Any = None,
    ) -> str:
        """Generate the itinerary prose with a single streaming LLM call.

        Writes rich Chinese Markdown directly from the *solved* itinerary, so the
        prose streams to the client immediately (real-time) with no blocking
        pre-pass. ``on_token`` (async ``(chunk)->None``) forwards each token to
        the frontend. Falls back to the plain skeleton on any failure.
        """
        skeleton = self._build_skeleton(itinerary, profile)
        if not skeleton:
            return ""
        try:
            from core.llm_client import llm

            # Output ≈ skeleton + per-POI prose; scale the budget to the input
            # (Chinese ≈ 1 token/char) so multi-day plans are never truncated.
            max_tokens = min(8192, max(2048, int(len(skeleton) * 2.5)))
            messages = [
                {
                    "role": "system",
                    "content": (
                        "你是专业的旅行行程文案撰写者。根据给定的行程数据，输出一份精美、"
                        "流畅的中文 Markdown 行程方案。要求：\n"
                        "1. 用 `#` 一级标题作为方案名（如“上海 5日游行程方案”）；\n"
                        "2. 紧接着一行给出预估费用（门票合计），并注明“不含住宿与往返大交通”；\n"
                        "3. 每天用 `##` 标题并配一个简短主题；\n"
                        "4. 每天的地点用有序列表：**加粗地点名**，保留时间段与门票价格，"
                        "换行后用 `_斜体_` 写一句吸引人的推荐理由；\n"
                        "5. 推荐理由可依据标签合理发挥，但**绝不可编造或修改地点名称、"
                        "时间、价格**；\n"
                        "6. 必须完整输出全部天数与全部地点，只返回 Markdown 文本。"
                    ),
                },
                {"role": "user", "content": skeleton},
            ]
            if on_token is not None:
                full = ""
                async for chunk in llm.stream_chat(
                    messages, temperature=0.5, max_tokens=max_tokens, task_type="output_format"
                ):
                    full += chunk
                    try:
                        await on_token(chunk)
                    except Exception:
                        pass  # token push is best-effort; never break generation
                return full.strip() or skeleton
            polished = await llm.chat(
                messages, temperature=0.5, max_tokens=max_tokens, task_type="output_format"
            )
            return polished.strip() or skeleton
        except Exception as exc:
            logger.warning("Streaming markdown generation failed: %s", exc)
            return skeleton

    async def build_artifacts(
        self,
        markdown: str,
        itinerary: list[dict[str, Any]] | None,
        city: str | None,
        session_id: str | None,
    ) -> dict[str, str | None]:
        """Build PDF / Excel / map links from finished prose + itinerary."""
        pdf_url = excel_url = None
        if settings.server_side_export_enabled:
            import asyncio

            file_id = session_id or uuid.uuid4().hex[:12]
            pdf_task = asyncio.create_task(self.generate_pdf(markdown, file_id))
            excel_task = asyncio.create_task(
                self.generate_excel(markdown, itinerary or [], file_id)
            )
            (_, pdf_url), (_, excel_url) = await asyncio.gather(pdf_task, excel_task)
        map_url = self.generate_map_url(itinerary, city)
        return {"pdf": pdf_url, "excel": excel_url, "map": map_url}

    async def polish_markdown(self, proposal_text: str, on_token: Any = None) -> str:
        """Return the itinerary Markdown, optionally LLM-polished.

        The writer (enrich) step already emits clean, fully-structured Markdown,
        so the extra LLM polish is disabled by default (``output_polish_enabled``).
        When disabled we stream the writer's prose directly — real content, paced
        for smooth client rendering, with no second ~2k-token generation. When
        enabled, fall back to the historical streaming LLM polish.
        """
        if not proposal_text:
            return ""

        if not settings.output_polish_enabled:
            await self._stream_existing(proposal_text, on_token)
            return proposal_text

        try:
            from core.llm_client import llm

            prompt_text = proposal_text[:12000]
            # The polished output reproduces the whole document, so its length
            # tracks the input. A fixed 2048-token cap truncated multi-day trips
            # mid-sentence; scale the budget to the prose length (Chinese ≈ 1
            # token/char) with headroom, capped at the model's 8K output limit.
            max_tokens = min(8192, max(2048, int(len(prompt_text) * 1.6)))
            messages = [
                {
                    "role": "system",
                    "content": (
                        "你是一位行程文档排版助手。请对以下 Markdown 行程进行排版和"
                        "润色（修正标题层级、统一列表符号、优化过渡语句），但**不要"
                        "修改任何景点名称、时间、价格、路线等事实信息**。必须完整输出"
                        "全部天数与景点，只返回 Markdown 文本。"
                    ),
                },
                {"role": "user", "content": prompt_text},
            ]
            if on_token is not None:
                full = ""
                async for chunk in llm.stream_chat(
                    messages, temperature=0.3, max_tokens=max_tokens, task_type="output_format"
                ):
                    full += chunk
                    try:
                        await on_token(chunk)
                    except Exception:
                        pass  # token push is best-effort; never break polishing
                return full.strip() or proposal_text
            polished = await llm.chat(
                messages, temperature=0.3, max_tokens=max_tokens, task_type="output_format"
            )
            return polished.strip() or proposal_text
        except Exception as exc:
            logger.warning("Markdown polish failed, using original: %s", exc)
            return proposal_text

    @staticmethod
    async def _stream_existing(text: str, on_token: Any) -> None:
        """Stream already-computed prose to the client in line-sized chunks.

        Gives a smooth, real progressive render without a second LLM call. Yields
        control between chunks so each is flushed to the SSE queue promptly; token
        pushes are best-effort and never raise.
        """
        if on_token is None:
            return
        import asyncio

        # Emit per line (keeping the newline) so Markdown structure renders
        # incrementally instead of arriving as one block.
        for line in text.splitlines(keepends=True):
            if not line:
                continue
            try:
                await on_token(line)
            except Exception:
                break  # client gone / queue full — stop, never break formatting
            await asyncio.sleep(0)  # flush this chunk before the next

    @staticmethod
    def _ensure_weasyprint_lib_path() -> None:
        """On macOS, WeasyPrint's CFFI loader needs Homebrew libs in dyld path."""
        if os.environ.get("DYLD_LIBRARY_PATH"):
            return
        if platform.system() != "Darwin":
            return
        for path in ("/opt/homebrew/lib", "/usr/local/lib"):
            if os.path.isdir(path):
                os.environ["DYLD_LIBRARY_PATH"] = path
                break

    async def generate_pdf(self, markdown_text: str, file_id: str) -> tuple[str | None, str | None]:
        """Convert Markdown to PDF via WeasyPrint if available."""
        if not markdown_text:
            return None, None
        self._ensure_weasyprint_lib_path()
        try:
            import markdown as md_lib
            from weasyprint import HTML
        except Exception as exc:
            logger.warning("PDF dependencies unavailable: %s", exc)
            return None, None

        try:
            html_body = md_lib.markdown(markdown_text)
            html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
body {{ font-family: "Noto Sans CJK SC", "PingFang SC", "Microsoft YaHei", sans-serif; margin: 2cm; }}
h1, h2, h3 {{ color: #2c3e50; }}
p {{ line-height: 1.6; }}
ul {{ margin-left: 1em; }}
</style>
</head>
<body>
{html_body}
</body>
</html>"""
            filename = _safe_filename(file_id, "pdf")
            path = os.path.join(PDF_DIR, filename)
            HTML(string=html).write_pdf(path)
            url = f"{self.base_url}/api/v1/download/pdfs/{filename}"
            return path, url
        except Exception as exc:
            logger.warning("PDF generation failed: %s", exc)
            return None, None

    async def generate_excel(
        self,
        markdown_text: str,
        itinerary: list[dict[str, Any]],
        file_id: str,
    ) -> tuple[str | None, str | None]:
        """Export day-by-day itinerary to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception as exc:
            logger.warning("openpyxl unavailable: %s", exc)
            return None, None

        try:
            wb = Workbook()
            ws = wb.active
            if ws is None:
                ws = wb.create_sheet("行程")
            ws.title = "行程"

            ws.append([" TravelAgent 行程单"])
            ws["A1"].font = Font(bold=True, size=14)
            ws.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")])
            ws.append([])
            ws.append(["天数", "时间", "类型", "名称", "时长(分钟)", "备注"])
            header_row = ws[4]
            for cell in header_row:
                cell.font = Font(bold=True)

            for day_idx, day in enumerate(itinerary, start=1):
                activities = day.get("activities") or day.get("items") or []
                for act in activities:
                    name = act.get("name") or act.get("poi_name") or ""
                    start_time = act.get("start_time") or act.get("time") or ""
                    category = act.get("category") or act.get("type") or ""
                    duration = act.get("duration_minutes") or act.get("duration") or ""
                    note = act.get("note") or act.get("reason") or ""
                    ws.append([f"第{day_idx}天", str(start_time), category, name, duration, note])

            # Summary sheet with plain-text Markdown.
            summary = wb.create_sheet("行程摘要")
            summary.append(["行程摘要"])
            for line in _strip_markdown_for_excel(markdown_text).splitlines():
                summary.append([line])

            filename = _safe_filename(file_id, "xlsx")
            path = os.path.join(EXCEL_DIR, filename)
            wb.save(path)
            url = f"{self.base_url}/api/v1/download/excel/{filename}"
            return path, url
        except Exception as exc:
            logger.warning("Excel generation failed: %s", exc)
            return None, None

    def generate_map_url(
        self, itinerary: list[dict[str, Any]] | None, city: str | None
    ) -> str | None:
        """Do not expose provider credentials in a client-visible URL.

        The frontend already renders ``TripMap`` from the real coordinates in
        the itinerary. Returning an AMap static URL previously leaked AMAP_KEY
        and plotted hash-derived fake coordinates, so no external artifact is
        produced here.
        """
        return None


# Singleton agent for the application.
output_format_agent = OutputFormatAgent()

__all__ = ["OutputFormatAgent", "output_format_agent", "PDF_DIR", "EXCEL_DIR"]
